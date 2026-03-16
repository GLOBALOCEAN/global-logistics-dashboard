```python
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# ────────────────────────────────────────────────
# Password Protection
# ────────────────────────────────────────────────
def check_password():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if st.session_state.authenticated:
        return True

    st.title("Login to Global Ocean Logistics Dashboard")

    username = st.text_input("Username")
    password = st.text_input("Password", type="password")

    if st.button("Login"):
        if username == "GLOBAL" and password == "Global123!":
            st.session_state.authenticated = True
            st.success("Logged in successfully!")
            st.rerun()
        else:
            st.error("Incorrect username or password")

    return False


if not check_password():
    st.stop()


# ────────────────────────────────────────────────
# Page Configuration
# ────────────────────────────────────────────────
st.set_page_config(
    page_title="Global Ocean Logistics Dashboard",
    page_icon="🌊",
    layout="wide"
)


# ────────────────────────────────────────────────
# Styling
# ────────────────────────────────────────────────
st.markdown("""
<style>
.stApp { background-color: #f8f9fa; }

h1, h2, h3 { color: #015486 !important; }

.stButton > button {
    background-color: #015486;
    color: white;
    border-radius: 6px;
    font-weight: bold;
}

.stButton > button:hover {
    background-color: #8fd8ff;
    color: #015486;
}
</style>
""", unsafe_allow_html=True)


# ────────────────────────────────────────────────
# Session State Navigation
# ────────────────────────────────────────────────
if "page" not in st.session_state:
    st.session_state.page = "home"


# ────────────────────────────────────────────────
# Sidebar
# ────────────────────────────────────────────────
with st.sidebar:

    st.markdown("<h2 style='color:#015486;'>Global Ocean Logistics</h2>", unsafe_allow_html=True)

    if st.button("🏠 Home", use_container_width=True):
        st.session_state.page = "home"
        st.rerun()

    if st.button("📊 Global Live Sheets", use_container_width=True):
        st.session_state.page = "live_sheets"
        st.rerun()

    if st.button("✈️ MAWB Tracker", use_container_width=True):
        st.session_state.page = "mawb"
        st.rerun()

    if st.button("📄 Generate Customer Tracker", use_container_width=True):
        st.session_state.page = "customer_tracker"
        st.rerun()


# ────────────────────────────────────────────────
# HOME PAGE
# ────────────────────────────────────────────────
if st.session_state.page == "home":

    st.title("Global Ocean Logistics Dashboard")
    st.markdown("Select a tool below")

    col1, col2, col3 = st.columns(3)

    with col1:
        if st.button("🌍 Global Live Sheets", use_container_width=True):
            st.session_state.page = "live_sheets"
            st.rerun()

    with col2:
        if st.button("✈️ MAWB Tracker", use_container_width=True):
            st.session_state.page = "mawb"
            st.rerun()

    with col3:
        if st.button("📄 Generate Customer Tracker", use_container_width=True):
            st.session_state.page = "customer_tracker"
            st.rerun()


# ────────────────────────────────────────────────
# LIVE SHEETS PAGE
# ────────────────────────────────────────────────
elif st.session_state.page == "live_sheets":

    st.title("Global Live Sheets")
    st.markdown("Open a tracker in a new tab")

    tracker_links = {
        "FCL Tracker":
        "https://netorgft11291460.sharepoint.com/:x:/r/sites/GlobalSharePoint/_layouts/15/Doc.aspx?sourcedoc=%7B27131CEF-29EA-4236-BF36-A3CF8D102D1D%7D",

        "LCL Tracker":
        "https://netorgft11291460.sharepoint.com/:x:/r/sites/GlobalSharePoint/_layouts/15/Doc.aspx?sourcedoc=%7B5AAFFF3D-9F8E-47E6-9755-8476A754408B%7D",

        "AIR Tracker":
        "https://netorgft11291460.sharepoint.com/:x:/r/sites/GlobalSharePoint/_layouts/15/Doc.aspx?sourcedoc=%7B587EDE6D-9415-42EF-9C11-01459CB389F5%7D"
    }

    selected = st.selectbox("Select Tracker", list(tracker_links.keys()))

    if st.button("Open Tracker", type="primary"):
        url = tracker_links[selected]
        st.markdown(f'<a href="{url}" target="_blank">Click here to open</a>', unsafe_allow_html=True)

    if st.button("← Back to Home"):
        st.session_state.page = "home"
        st.rerun()


# ────────────────────────────────────────────────
# MAWB PAGE (placeholder)
# ────────────────────────────────────────────────
elif st.session_state.page == "mawb":

    st.title("MAWB Tracker")
    st.info("Insert your MAWB tracker code here")

    if st.button("← Back to Home"):
        st.session_state.page = "home"
        st.rerun()


# ────────────────────────────────────────────────
# CUSTOMER TRACKER PAGE
# ────────────────────────────────────────────────
elif st.session_state.page == "customer_tracker":

    st.title("Generate Customer-Specific Tracker")

    st.markdown("""
Upload your **FCL or LCL tracker file**, choose the **customer**, and download a **filtered tracker**.
""")

    uploaded_file = st.file_uploader(
        "Upload Tracker Excel File",
        type=["xlsx"]
    )

    if uploaded_file:

        df = pd.read_excel(uploaded_file)
        df.columns = df.columns.str.strip()

        st.subheader("Preview")
        st.dataframe(df.head())

        consignee_col = st.selectbox(
            "Select Consignee / Customer Column",
            df.columns
        )

        customers = sorted(df[consignee_col].dropna().astype(str).unique())

        customer_name = st.selectbox(
            "Select Customer",
            customers
        )

        tracker_type = st.selectbox(
            "Tracker Type",
            ["FCL", "LCL"]
        )

        if st.button("Generate & Download Tracker", type="primary"):

            filtered = df[
                df[consignee_col]
                .astype(str)
                .str.contains(customer_name, case=False, na=False)
            ]

            if filtered.empty:
                st.error("No rows found")

            else:

                wb = Workbook()
                ws = wb.active

                title = f"MY LIFE BATHROOM {tracker_type} FREIGHT TRACKER"

                last_col = get_column_letter(len(filtered.columns))
                ws.merge_cells(f"A1:{last_col}1")

                cell = ws["A1"]
                cell.value = title
                cell.font = Font(bold=True, size=16, color="FFFFFF")
                cell.fill = PatternFill(start_color="005566", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

                headers = list(filtered.columns)

                for col_num, header in enumerate(headers, 1):
                    ws.cell(row=2, column=col_num, value=header)

                for row_num, row_data in enumerate(filtered.itertuples(index=False), 3):
                    for col_num, value in enumerate(row_data, 1):
                        ws.cell(row=row_num, column=col_num, value=value)

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                today = datetime.now().strftime("%Y-%m-%d")
                safe_name = customer_name.replace(" ", "_")

                filename = f"{tracker_type}_Tracker_{safe_name}_{today}.xlsx"

                st.download_button(
                    label="Download Tracker",
                    data=output,
                    file_name=filename
                )

    if st.button("← Back to Home"):
        st.session_state.page = "home"
        st.rerun()
```
