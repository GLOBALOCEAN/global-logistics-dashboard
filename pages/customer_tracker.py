import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

st.title("📄 Multi-Tracker Customer Report")

# ────────────────────────────────────────────────
# FILE UPLOADS
# ────────────────────────────────────────────────
fcl_file = st.file_uploader("Upload FCL Tracker", type=["xlsx"])
lcl_file = st.file_uploader("Upload LCL Tracker", type=["xlsx"])
air_file = st.file_uploader("Upload AIR Tracker", type=["xlsx"])

dfs = {}

if fcl_file:
    dfs["FCL"] = pd.read_excel(fcl_file)

if lcl_file:
    dfs["LCL"] = pd.read_excel(lcl_file)

if air_file:
    dfs["AIR"] = pd.read_excel(air_file)

# ────────────────────────────────────────────────
# MAIN LOGIC
# ────────────────────────────────────────────────
if dfs:

    # Clean column names
    for key in dfs:
        dfs[key].columns = dfs[key].columns.str.strip()

    sample_df = list(dfs.values())[0]

    # Select customer column
    customer_column = st.selectbox(
        "Select Customer Column",
        sample_df.columns
    )

    # ────────────────────────────────────────────────
    # BUILD CLEAN CUSTOMER LIST (FIXED)
    # ────────────────────────────────────────────────
    all_customers = set()

    for df in dfs.values():
        if customer_column in df.columns:
            cleaned = (
                df[customer_column]
                .astype(str)
                .str.replace("\xa0", " ", regex=False)
                .str.replace("\n", " ", regex=False)
                .str.strip()
            )

            # Remove junk values
            cleaned = cleaned[cleaned != ""]
            cleaned = cleaned[cleaned.str.lower() != "nan"]

            all_customers.update(cleaned.tolist())

    # Force clean sorted list
    customer_list = sorted([str(c) for c in all_customers])

    selected_customer = st.selectbox(
        "Select Customer",
        customer_list
    )

    # ────────────────────────────────────────────────
    # STATUS FILTER (FCL/LCL ONLY)
    # ────────────────────────────────────────────────
    status_options = []

    for df in dfs.values():
        if "Status" in df.columns:
            status_options.extend(
                df["Status"]
                .astype(str)
                .str.strip()
                .unique()
            )

    status_options = sorted(set(status_options))

    selected_status = st.multiselect(
        "Filter by Status (FCL/LCL only)",
        options=status_options,
        default=status_options
    )

    # ────────────────────────────────────────────────
    # GENERATE REPORT
    # ────────────────────────────────────────────────
    if st.button("Generate Multi-Tracker Report", type="primary"):

        wb = Workbook()
        wb.remove(wb.active)

        customer_clean = selected_customer.lower().strip()

        for name, df in dfs.items():

            if customer_column not in df.columns:
                st.warning(f"{name}: Missing '{customer_column}' column")
                continue

            # Clean customer column
            df[customer_column] = (
                df[customer_column]
                .astype(str)
                .str.replace("\xa0", " ", regex=False)
                .str.replace("\n", " ", regex=False)
                .str.strip()
                .str.lower()
            )

            # Filter customer
            filtered = df[
                df[customer_column].str.contains(customer_clean, na=False)
            ]

            # Apply status filter if exists
            if "Status" in df.columns:
                df["Status"] = df["Status"].astype(str).str.strip()
                filtered = filtered[filtered["Status"].isin(selected_status)]

            if filtered.empty:
                st.warning(f"{name}: No matches found for '{selected_customer}'")
                continue

            # ────────────────────────────────────────────────
            # CREATE SHEET
            # ────────────────────────────────────────────────
            ws = wb.create_sheet(title=name)

            # TITLE
            title = f"GLOBAL OCEAN LOGISTICS - {name} SHIPMENT TRACKER"
            last_col = get_column_letter(len(filtered.columns))

            ws.merge_cells(f"A1:{last_col}1")
            cell = ws["A1"]
            cell.value = title
            cell.font = Font(bold=True, size=16, color="FFFFFF")
            cell.fill = PatternFill(start_color="005566", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

            # INFO ROWS
            ws["A2"] = f"Customer: {selected_customer}"
            ws["A3"] = f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}"

            # HEADERS
            headers = list(filtered.columns)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="005566", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # DATA
            status_index = None
            if "Status" in filtered.columns:
                status_index = filtered.columns.get_loc("Status")

            for row_num, row_data in enumerate(filtered.itertuples(index=False), 6):

                status_value = ""
                if status_index is not None:
                    status_value = str(row_data[status_index]).strip()

                for col_num, value in enumerate(row_data, 1):

                    cell = ws.cell(row=row_num, column=col_num, value=value)

                    # Colour coding
                    if status_index is not None:

                        if status_value == "In Transit":
                            cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")

                        elif status_value == "Waiting to Sail":
                            cell.fill = PatternFill(start_color="FFEB9C", fill_type="solid")

                        elif status_value == "Awaiting Confirmation":
                            cell.fill = PatternFill(start_color="BDD7EE", fill_type="solid")

                        elif status_value == "Arrived":
                            cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                    cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # AUTO WIDTH
            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)

                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[column_letter].width = max_length + 3

        # FINAL OUTPUT
        if not wb.sheetnames:
            st.error("No data found in ANY tracker for this customer.")
        else:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            filename = f"{selected_customer}_Full_Shipment_Report_{datetime.now().strftime('%d-%m-%Y')}.xlsx"

            st.download_button(
                "Download Report",
                data=output,
                file_name=filename
            )
